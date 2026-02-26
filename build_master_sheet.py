"""
build_master_sheet.py
---------------------
One-time script to:
1. Scrape all schools from WI_School_List.xlsx
2. Build a master Google Sheet with Schools + Contacts tabs
3. Populate all school info and contacts with Sync = N by default

Run once to initialize, then ns_school_sync.py handles daily updates.
"""

import requests
import time
import re
import openpyxl
from bs4 import BeautifulSoup
from netsuite_sync import (scrape_wiaa_school_detail, slugify,
                            WIAA_HEADERS, NAV_H5S)

INPUT_FILE  = "WI School List.xlsx"
OUTPUT_FILE = "School_Master_List.xlsx"

# ============================================================
# LOAD SCHOOL LIST
# ============================================================
wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in['Schools']
schools = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:
        schools.append({"name": row[0], "url": row[1]})

print(f"Found {len(schools)} schools to scrape\n")

# ============================================================
# SCRAPE ALL SCHOOLS
# ============================================================
all_school_rows = []
all_contact_rows = []

for i, s in enumerate(schools):
    print(f"[{i+1}/{len(schools)}] {s['name']}...")
    info, admins, coaches = scrape_wiaa_school_detail(s["url"])

    level = info.get("level", "")
    full_name = f"{s['name']} {level}".strip() if level else s["name"]

    all_school_rows.append({
        "school_name":   s["name"],
        "full_name":     full_name,
        "state":         info.get("state", "WI"),
        "url":           s["url"],
        "ns_ext_id":     slugify(s["name"]),
        "sales_rep":     "",
        "school_class":  info.get("school_class", ""),
        "level":         level,
        "nickname":      info.get("nickname", ""),
        "colors":        info.get("colors", ""),
        "conference":    info.get("conference", ""),
        "wiaa_district": info.get("wiaa_district", ""),
        "enrollment":    info.get("enrollment", ""),
        "school_size":   info.get("school_size", ""),
        "phone":         info.get("phone", ""),
        "address1":      info.get("address1", ""),
        "address2":      info.get("address2", ""),
        "city":          info.get("city", ""),
        "zip":           info.get("zip", ""),
        "website":       info.get("website", ""),
        "last_synced":   "",
        "ns_customer_id": "",
        "notes":         "",
    })

    for c in admins + coaches:
        all_contact_rows.append({
            "school_name": s["name"],
            "full_name":   full_name,
            "first":       c.get("first", ""),
            "last":        c.get("last", ""),
            "email":       c.get("email", ""),
            "role":        c.get("role", ""),
            "type":        c.get("type", ""),
            "sync":        "N",
            "ns_contact_id": "",
            "last_synced": "",
        })

    print(f"  -> {len(admins)} admins, {len(coaches)} coaches")
    time.sleep(1)  # be polite to WIAA server

# ============================================================
# BUILD EXCEL FILE
# ============================================================
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- SCHOOLS TAB ----
ws_schools = wb.active
ws_schools.title = "Schools"

school_headers = [
    "School Name", "Full Name", "State", "School URL", "NS Ext ID",
    "Sales Rep", "Class", "Level", "Nickname", "Colors", "Conference",
    "WIAA District", "Enrollment", "Size", "Phone",
    "Address1", "Address2", "City", "Zip", "Website",
    "NS Customer ID", "Last Synced", "Notes"
]

header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF")

for col, h in enumerate(school_headers, 1):
    cell = ws_schools.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, s in enumerate(all_school_rows, 2):
    ws_schools.cell(row=row_idx, column=1,  value=s["school_name"])
    ws_schools.cell(row=row_idx, column=2,  value=s["full_name"])
    ws_schools.cell(row=row_idx, column=3,  value=s["state"])
    ws_schools.cell(row=row_idx, column=4,  value=s["url"])
    ws_schools.cell(row=row_idx, column=5,  value=s["ns_ext_id"])
    ws_schools.cell(row=row_idx, column=6,  value=s["sales_rep"])
    ws_schools.cell(row=row_idx, column=7,  value=s["school_class"])
    ws_schools.cell(row=row_idx, column=8,  value=s["level"])
    ws_schools.cell(row=row_idx, column=9,  value=s["nickname"])
    ws_schools.cell(row=row_idx, column=10, value=s["colors"])
    ws_schools.cell(row=row_idx, column=11, value=s["conference"])
    ws_schools.cell(row=row_idx, column=12, value=s["wiaa_district"])
    ws_schools.cell(row=row_idx, column=13, value=s["enrollment"])
    ws_schools.cell(row=row_idx, column=14, value=s["school_size"])
    ws_schools.cell(row=row_idx, column=15, value=s["phone"])
    ws_schools.cell(row=row_idx, column=16, value=s["address1"])
    ws_schools.cell(row=row_idx, column=17, value=s["address2"])
    ws_schools.cell(row=row_idx, column=18, value=s["city"])
    ws_schools.cell(row=row_idx, column=19, value=s["zip"])
    ws_schools.cell(row=row_idx, column=20, value=s["website"])
    ws_schools.cell(row=row_idx, column=21, value=s["ns_customer_id"])
    ws_schools.cell(row=row_idx, column=22, value=s["last_synced"])
    ws_schools.cell(row=row_idx, column=23, value=s["notes"])
    # Alternate row shading
    if row_idx % 2 == 0:
        for col in range(1, 24):
            ws_schools.cell(row=row_idx, column=col).fill = \
                PatternFill("solid", fgColor="DCE6F1")

# Column widths
col_widths = [20,30,8,60,25,12,12,15,15,18,18,18,12,10,16,30,15,15,10,50,16,18,20]
for i, w in enumerate(col_widths, 1):
    ws_schools.column_dimensions[get_column_letter(i)].width = w
ws_schools.freeze_panes = "A2"

# ---- CONTACTS TAB ----
ws_contacts = wb.create_sheet("Contacts")

contact_headers = [
    "School Name", "Full School Name", "First Name", "Last Name",
    "Email", "Role", "Type", "Sync (Y/N)",
    "NS Contact ID", "Last Synced"
]

for col, h in enumerate(contact_headers, 1):
    cell = ws_contacts.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, c in enumerate(all_contact_rows, 2):
    ws_contacts.cell(row=row_idx, column=1,  value=c["school_name"])
    ws_contacts.cell(row=row_idx, column=2,  value=c["full_name"])
    ws_contacts.cell(row=row_idx, column=3,  value=c["first"])
    ws_contacts.cell(row=row_idx, column=4,  value=c["last"])
    ws_contacts.cell(row=row_idx, column=5,  value=c["email"])
    ws_contacts.cell(row=row_idx, column=6,  value=c["role"])
    ws_contacts.cell(row=row_idx, column=7,  value=c["type"])
    ws_contacts.cell(row=row_idx, column=8,  value=c["sync"])
    ws_contacts.cell(row=row_idx, column=9,  value=c["ns_contact_id"])
    ws_contacts.cell(row=row_idx, column=10, value=c["last_synced"])
    if row_idx % 2 == 0:
        for col in range(1, 11):
            ws_contacts.cell(row=row_idx, column=col).fill = \
                PatternFill("solid", fgColor="DCE6F1")

contact_widths = [20, 30, 15, 20, 40, 35, 10, 12, 16, 18]
for i, w in enumerate(contact_widths, 1):
    ws_contacts.column_dimensions[get_column_letter(i)].width = w
ws_contacts.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"\n✅ Saved: {OUTPUT_FILE}")
print(f"   Schools:  {len(all_school_rows)}")
print(f"   Contacts: {len(all_contact_rows)}")
