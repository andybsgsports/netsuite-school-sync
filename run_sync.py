"""
run_sync.py
-----------
Master runner. Reads School_Master_List.xlsx and syncs all schools to NetSuite.
- Looks up NS Customer ID from NetSuite via Ext ID if not already in spreadsheet
- Updates both child and parent customer records when a parent relationship exists
- Writes NS Customer ID back to spreadsheet after first successful lookup
Run directly: python run_sync.py
Scheduled:    Task Scheduler -> run_sync.bat -> this file
"""

import openpyxl
import time
from datetime import datetime
from netsuite_sync import sync_school, sync_parent_record, get_customer_by_external_id, ns_get

EXCEL_FILE    = "School Sync Master.xlsx"
SCHOOL_FILTER = ""  # Set to a school name to test one. Blank = all.

def load_sheet():
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # ---- Schools tab ----
    ws = wb["Schools"]
    rows = list(ws.iter_rows(values_only=True))
    hdr  = list(rows[0])

    def col(name):
        return hdr.index(name) if name in hdr else -1

    # Add NS Parent Customer ID column if missing
    if "NS Parent Customer ID" not in hdr:
        next_col = len(hdr) + 1
        ws.cell(row=1, column=next_col, value="NS Parent Customer ID")
        hdr.append("NS Parent Customer ID")

    S = {
        "name":       col("School Name"),
        "url":        col("School URL"),
        "state":      col("State"),
        "sales_rep":  col("Sales Rep"),
        "ext_id":     col("NS Ext ID"),
        "ns_id":      col("NS Customer ID"),
        "parent_id":  col("NS Parent Customer ID"),
        "last_synced":col("Last Synced"),
    }

    schools = []
    for row_idx, row in enumerate(rows[1:], start=2):
        name = (row[S["name"]] or "").strip() if S["name"] > -1 else ""
        url  = (row[S["url"]]  or "").strip() if S["url"]  > -1 else ""
        if not name or not url:
            continue
        schools.append({
            "name":      name,
            "url":       url,
            "state":     (row[S["state"]]    or "WI").strip() if S["state"]    > -1 else "WI",
            "sales_rep": (row[S["sales_rep"]] or "").strip()  if S["sales_rep"]> -1 else "",
            "ext_id":    (row[S["ext_id"]]    or "").strip()  if S["ext_id"]   > -1 else "",
            "ns_id":     str(row[S["ns_id"]]  or "").strip()  if S["ns_id"]    > -1 else "",
            "parent_id": str(row[S["parent_id"]] or "").strip() if S["parent_id"] > -1 and S["parent_id"] < len(row) else "",
            "row":       row_idx,
        })

    # ---- Contacts tab ----
    ws2   = wb["Contacts"]
    rows2 = list(ws2.iter_rows(values_only=True))
    hdr2  = list(rows2[0])

    def col2(name):
        return hdr2.index(name) if name in hdr2 else -1

    C = {
        "school": col2("School Name"),
        "first":  col2("First Name"),
        "last":   col2("Last Name"),
        "email":  col2("Email"),
        "role":   col2("Role"),
        "type":   col2("Type"),
        "sync":   col2("Sync (Y/N)"),
        "ns_id":  col2("NS Contact ID"),
    }

    contacts_by_school = {}
    for row in rows2[1:]:
        school = (row[C["school"]] or "").strip() if C["school"] > -1 else ""
        if not school:
            continue
        sync_flag = (row[C["sync"]] or "").strip().upper() if C["sync"] > -1 else "N"
        if school not in contacts_by_school:
            contacts_by_school[school] = []
        contacts_by_school[school].append({
            "first":  (row[C["first"]]  or "").strip() if C["first"]  > -1 else "",
            "last":   (row[C["last"]]   or "").strip() if C["last"]   > -1 else "",
            "email":  (row[C["email"]]  or "").strip() if C["email"]  > -1 else "",
            "role":   (row[C["role"]]   or "").strip() if C["role"]   > -1 else "",
            "type":   (row[C["type"]]   or "").strip() if C["type"]   > -1 else "",
            "sync":   sync_flag == "Y",
        })

    return schools, contacts_by_school, wb, ws, S

def get_ns_id(school):
    """Return NS Customer internal ID, looking it up via Ext ID if needed."""
    if school["ns_id"]:
        return school["ns_id"]
    if school["ext_id"]:
        ns_id = get_customer_by_external_id(school["ext_id"])
        if ns_id:
            print(f"  [NS] Found via Ext ID '{school['ext_id']}': {ns_id}")
            return str(ns_id)
    print(f"  [SKIP] Cannot find NS Customer ID for {school['name']}")
    return None

def get_parent_id(ns_id):
    """Check if this customer has a parent in NetSuite. Returns parent ID or None."""
    r = ns_get(f"customer/{ns_id}?fields=parent")
    if r.status_code == 200:
        parent = r.json().get("parent")
        if parent:
            return str(parent.get("id", ""))
    return None

def main():
    print(f"\n{'='*60}")
    print(f"  BADGER SCHOOL SYNC  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}")

    schools, contacts_by_school, wb, ws, S = load_sheet()

    if SCHOOL_FILTER:
        schools = [s for s in schools if s["name"] == SCHOOL_FILTER]
        print(f"  TEST MODE: Only syncing {SCHOOL_FILTER}")

    print(f"  Schools to sync: {len(schools)}\n")

    synced = 0
    skipped = 0
    errors = 0

    for school in schools:
        name      = school["name"]
        url       = school["url"]
        state     = school["state"]
        sales_rep = school["sales_rep"]

        # Get or look up NS Customer ID
        ns_id = get_ns_id(school)
        if not ns_id:
            skipped += 1
            continue

        # Write NS Customer ID back to spreadsheet if it was missing
        if not school["ns_id"] and S["ns_id"] > -1:
            ws.cell(row=school["row"], column=S["ns_id"] + 1, value=ns_id)

        # Check for parent relationship
        parent_id = school["parent_id"] or get_parent_id(ns_id)
        if parent_id and not school["parent_id"] and S["parent_id"] > -1:
            ws.cell(row=school["row"], column=S["parent_id"] + 1, value=parent_id)

        school_contacts = contacts_by_school.get(name, [])

        try:
            # Sync the school (child) record
            result_id, school_info, all_found, created = sync_school(
                school_name    = name,
                school_url     = url,
                state          = state,
                sync_contacts  = school_contacts,
                sales_rep      = sales_rep,
                ns_customer_id = ns_id,
            )

            # If parent exists, sync contacts and addresses ONLY (do not overwrite district data)
            if result_id and parent_id:
                contacts_to_sync = [c for c in school_contacts if c.get("sync")]
                sync_parent_record(parent_id, school_info, contacts_to_sync)

            if result_id:
                synced += 1
                if S["last_synced"] > -1:
                    ws.cell(row=school["row"], column=S["last_synced"] + 1,
                            value=datetime.now().strftime("%Y-%m-%d %H:%M"))
            else:
                errors += 1

        except Exception as e:
            print(f"  [ERROR] {name}: {e}")
            errors += 1

        time.sleep(1)

    wb.save(EXCEL_FILE)

    print(f"\n{'='*60}")
    print(f"  SYNC COMPLETE")
    print(f"  Synced: {synced}  |  Skipped: {skipped}  |  Errors: {errors}")
    print(f"  Finished: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
