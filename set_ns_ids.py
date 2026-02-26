"""
set_ns_ids.py
-------------
Directly writes the correct NS Customer IDs into School_Master_List.xlsx.
Run once: python set_ns_ids.py
"""
import openpyxl

EXCEL_FILE = "School_Master_List.xlsx"

# School Name -> correct NS Internal ID
# IDs confirmed from populate_ns_ids.py output — picking the actual High School record
CORRECT_IDS = {
    "Barneveld":                  "994",    # Barneveld School District
    "Beloit Memorial":            "1029",   # BELOIT MEMORIAL HIGH SCHOOL
    "Big Foot":                   "1009",   # BIG FOOT UHS SCHOOL DIST
    "Brookfield Academy":         "667",    # BROOKFIELD ACADEMY HIGH SCHOOL
    "Brookfield Central":         "1094",   # BROOKFIELD CENTRAL HIGH SCHOOL
    "Brookfield East":            "1037",   # BROOKFIELD EAST HIGH SCH
    "Brown Deer":                 "1047",   # BROWN DEER HIGH SCHOOL
    "Burlington":                 "1048",   # Burlington High School
    "Catholic Central - Burlington": "1040", # BURLINGTON CATHOLIC CENTRAL
    "Clinton":                    "1265",   # CLINTON HIGH SCHOOL
    "Delavan-Darien":             "1436",   # DELAVAN-DARIEN S.D.
    "East Troy":                  "1562",   # EAST TROY HIGH SCHOOL
    "Elkhorn":                    "8721",   # Created 2/25/2026
    "Janesville Craig":           "1967",   # JANESVILLE CRAIG H.S.
    "Janesville Parker":          "1968",   # Janesville Parker High School
    "Kenosha Bradford":           "8727",   # Created 2/25/2026
    "Kenosha St. Joseph Catholic":"8728",   # Created 2/25/2026
    "Kenosha Tremper":            "2033",   # KENOSHA TREMPER HIGH SCH
    "Mount Horeb":                "2217",   # Mt. Horeb High School
    "Mukwonago":                  "2318",   # MUKWONAGO HIGH SCHOOL
    "Palmyra-Eagle":              "2849",   # Palmyra-Eagle High School
    "Racine Lutheran":            "3000",   # RACINE LUTHERAN HIGH SCH
    "Racine St. Catherine's":     "8736",   # Created 2/25/2026
    "Beloit Turner":              "1002",   # Beloit Turner School District
    "Union Grove":                "3551",   # UNION GROVE UNION HIGH SCHOOL
    "Verona Area":                "3589",   # Verona High School
    "Waterford":                  "3651",   # Waterford Union High School
    "Westosha Central":           "1230",   # WESTOSHA CENTRAL HIGH SCHOOL DISTRICT
    "Whitnall":                   "3666",   # Whitnall High School
    "Whitewater":                 "3670",   # Whitewater High School
    "Williams Bay":               "3710",   # Williams Bay School District
    "Wilmot Union":               "3711",   # WILMOT UNION HIGH SCHOOL
    # Newly created 2/25/2026 morning run
    "Greendale":                  "8722",
    "Greenfield":                 "8723",
    "Hamilton":                   "8724",
    "Menomonee Falls":            "8730",
    "Oconomowoc":                 "8733",
    "Waukesha North":             "8741",
    "Waukesha South":             "8742",
    "Waukesha West":              "8743",
    "Wauwatosa West":             "8744",
    "West Bend East":             "8745",
    "West Bend West":             "8746",
}

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb["Schools"]
rows = list(ws.iter_rows(values_only=True))
hdr  = list(rows[0])

name_col = hdr.index("School Name")
id_col   = hdr.index("NS Customer ID")

updated = 0
for row_idx, row in enumerate(rows[1:], start=2):
    school_name = (row[name_col] or "").strip()
    if school_name in CORRECT_IDS:
        ns_id = CORRECT_IDS[school_name]
        if ns_id:
            ws.cell(row=row_idx, column=id_col + 1, value=ns_id)
            print(f"  [SET] {school_name} = {ns_id}")
            updated += 1
        else:
            ws.cell(row=row_idx, column=id_col + 1, value=None)
            print(f"  [BLANK] {school_name} - needs manual lookup")

wb.save(EXCEL_FILE)
print(f"\nDone! {updated} IDs written.")
