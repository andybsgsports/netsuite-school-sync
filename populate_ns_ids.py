"""
populate_ns_ids.py
------------------
One-time script. Reads CustomersProjects412.csv and matches school names
to School_Master_List.xlsx, writing NS Customer IDs into the NS Customer ID column.

Run once:
  python populate_ns_ids.py

Then run_sync.py will have the IDs it needs.
"""

import csv
import openpyxl

CSV_FILE   = "CustomersProjects412.csv"
EXCEL_FILE = "School_Master_List.xlsx"

def norm(s):
    s = s.lower().strip()
    for w in ["high school", "hs", "school district", "district", "school"]:
        s = s.replace(w, "")
    s = (s.replace("st.", "saint")
          .replace("mt.", "mount")
          .replace("-", " ")
          .replace("'", ""))
    return " ".join(s.split())

# Load CSV
print(f"Loading {CSV_FILE}...")
ns_customers = []
with open(CSV_FILE, encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        ns_customers.append({
            "id":   row["Internal ID"],
            "name": row["Company Name"],
            "norm": norm(row["Company Name"]),
        })
print(f"  {len(ns_customers)} customers loaded")

# Load Excel
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb["Schools"]
rows = list(ws.iter_rows(values_only=True))
hdr  = list(rows[0])

name_col = hdr.index("School Name")
id_col   = hdr.index("NS Customer ID")

print(f"\nMatching schools...\n")
updated = 0
for row_idx, row in enumerate(rows[1:], start=2):
    school_name = (row[name_col] or "").strip()
    current_id  = str(row[id_col] or "").strip()

    if not school_name:
        continue
    if current_id:
        print(f"  [SKIP]  {school_name} already has ID: {current_id}")
        continue

    nn   = norm(school_name)
    hits = [c for c in ns_customers if nn in c["norm"] or c["norm"] in nn]

    if len(hits) == 1:
        print(f"  [FOUND] {school_name} -> {hits[0]['id']} ({hits[0]['name']})")
        ws.cell(row=row_idx, column=id_col + 1, value=hits[0]["id"])
        updated += 1
    elif len(hits) > 1:
        print(f"  [MULTI] {school_name} — {len(hits)} matches, pick one:")
        for h in hits:
            print(f"           {h['id']}  {h['name']}")
    else:
        print(f"  [MISS]  {school_name} [{nn}]")

wb.save(EXCEL_FILE)
print(f"\nDone! {updated} IDs written to {EXCEL_FILE}")
