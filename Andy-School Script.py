# Andy-School Script.py
# ------------------------------------------------------------
# Build a single WIAA-style workbook that includes:
#   • Fresh WIAA scrape (from WI School List.xlsx)
#   • IHSA results loaded from IHSA-Batch-Output\schools\IHSA-Batch-Combined.xlsx
# Output keeps WIAA formatting/sheets:
#   - "Athletic Admins" (AD-only per rules below)
#   - "Administrators" (others; Activities Director & Assistant Athletic Director stay here)
#   - One sheet per Sport Group for coaches/activities
#   - "State" column on every sheet (ALL CAPS)
# NetSuite sync: diffs are pushed to NetSuite Customer/Contact records
# ------------------------------------------------------------

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import re
import shutil
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys
import io
from pathlib import Path
from glob import glob
from typing import List

# NetSuite sync module
from netsuite_sync import sync_changes_to_netsuite

# Force UTF-8 output even when redirected or run from Task Scheduler
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============= BEGIN PATH CONFIG =============
print(f"[RUNNING SCRIPT] {__file__}")

BASE_DIR = r"C:\Users\andre\OneDrive - Badger Sporting Goods\Desktop\Illinois Contact List"

EXCEL_INPUT = os.path.join(BASE_DIR, "WI School List.xlsx")
SHEET_NAME  = "Schools"

FINAL_XLSX  = os.path.join(BASE_DIR, "Andy-School Admins and Coaches.xlsx")
PREV_XLSX   = os.path.join(BASE_DIR, "Andy-Previous School Admins and Coaches.xlsx")

IHSA_COMBINED_XLSX = os.path.join(BASE_DIR, "IHSA-Batch-Output", "schools", "IHSA-Batch-Combined.xlsx")

# Email settings
SEND_EMAIL   = True
EMAIL_TO     = "andy@bsgsports.com"
EMAIL_BCC    = ""
EMAIL_SUBJ   = "Updated WI+IL School Admins and Coaches"

# NetSuite sync setting - set to False to disable NS sync
SYNC_TO_NETSUITE = True

if not os.path.exists(EXCEL_INPUT):
    candidates = glob(os.path.join(BASE_DIR, "*School*List*.xlsx"))
    if candidates:
        EXCEL_INPUT = candidates[0]
        print(f"[INFO] Using detected input workbook: {EXCEL_INPUT}")
    else:
        raise FileNotFoundError(
            f"Input workbook not found. Looked for 'WI School List.xlsx' in {BASE_DIR}."
        )

print(f"[CONFIG] BASE_DIR={BASE_DIR}")
print(f"[CONFIG] EXCEL_INPUT={EXCEL_INPUT}")
print(f"[CONFIG] IHSA_COMBINED_XLSX={IHSA_COMBINED_XLSX}")
print(f"[CONFIG] FINAL_XLSX={FINAL_XLSX}")
# ============= END PATH CONFIG =============

ATHLETIC_AD_SHEET_ROLES = {
    "Athletic Director",
    "Assistant Principal, Athletic Director",
    "Boys Athletic Director",
    "Girls Athletic Director",
}

# ---------- Helpers ----------
def decode_email(encoded_email):
    r = int(encoded_email[:2], 16)
    return "".join(chr(int(encoded_email[i:i+2], 16) ^ r) for i in range(2, len(encoded_email), 2))

def split_name(full_name):
    clean = full_name.replace("\u00a0", " ").replace("\xa0", " ").strip()
    parts = clean.split(" ", 1)
    if len(parts) == 2:
        return parts[0].title(), parts[1].title()
    return clean.title(), ""

def _norm(s):  return re.sub(r"\s+"," ", ("" if s is None else str(s)).strip())
def _title(s): return _norm(s).title()

def state_code(val, default="IL"):
    s = (val or default)
    return re.sub(r"[^A-Za-z]", "", str(s)).upper()[:2] or default

def combine_sports(sports_list: List[str]):
    if not sports_list or len(sports_list) == 1:
        return sports_list[0] if sports_list else ""
    unique_sports = list(dict.fromkeys(sports_list))
    if len(unique_sports) == 1:
        return unique_sports[0]
    has_boys  = any("boys"  in s.lower() for s in unique_sports)
    has_girls = any("girls" in s.lower() for s in unique_sports)
    if has_boys and has_girls:
        base = re.sub(r'\b(Boys|Girls)\b', '', unique_sports[0], flags=re.IGNORECASE).strip()
        return f"Boys & Girls {base}"
    return " & ".join(unique_sports)

def combine_coaching_roles(roles_list, sports_list):
    if not roles_list or len(roles_list) == 1:
        return roles_list[0] if roles_list else "", sports_list[0] if sports_list else ""
    types = set()
    for role in roles_list:
        rl = (role or "").lower()
        if "head coach" in rl: types.add("Head Coach")
        elif "assistant coach" in rl: types.add("Assistant Coach")
        elif "coach" in rl: types.add("Coach")
        else: types.add(role)
    combined_sport = combine_sports(sports_list)
    if "Head Coach" in types: final_role = "Head Coach"
    elif "Assistant Coach" in types: final_role = "Assistant Coach"
    elif "Coach" in types: final_role = "Coach"
    else: final_role = " & ".join(sorted(types))
    return final_role, combined_sport

def deduplicate_coaches_by_school_email(df):
    if df.empty: return df
    out = []
    for (school, email), g in df.groupby(['School', 'Email']):
        if len(g) == 1:
            out.append(g.iloc[0].to_dict())
        else:
            first = g.iloc[0].to_dict()
            r, s = combine_coaching_roles(g['Role'].tolist(), g['Sport'].tolist())
            first['Role']  = r
            first['Sport'] = s
            out.append(first)
    return pd.DataFrame(out)

def canonical_admin_role_preserve_gender(role: str) -> str:
    r = _norm(role)
    low = r.lower()
    if ("athletic director" in low and "assistant" not in low
            and "boys" in low and "girls" in low):
        return "Athletic Director"
    if "assistant principal" in low and "athletic director" in low:
        return "Assistant Principal, Athletic Director"
    if "assistant athletic director" in low:
        return "Assistant Athletic Director"
    if "activities director" in low:
        return "Activities Director"
    if "supervisor" in low:
        return _title(r)
    if "athletic director" in low and "assistant" not in low:
        if "boys" in low:  return "Boys Athletic Director"
        if "girls" in low: return "Girls Athletic Director"
        return "Athletic Director"
    return _title(r)

def collapse_boys_girls_ad_per_person(roles: list[str]) -> str:
    canon = [canonical_admin_role_preserve_gender(r) for r in roles if r]
    s = set(canon)
    if "Assistant Principal, Athletic Director" in s:
        return "Assistant Principal, Athletic Director"
    has_plain_ad   = "Athletic Director" in s
    has_boys_ad    = "Boys Athletic Director" in s
    has_girls_ad   = "Girls Athletic Director" in s
    if (has_boys_ad and has_girls_ad) or has_plain_ad:
        return "Athletic Director"
    if has_boys_ad:  return "Boys Athletic Director"
    if has_girls_ad: return "Girls Athletic Director"
    ordered = sorted(s)
    return " & ".join(ordered)

def deduplicate_admins_by_school_email(df):
    if df.empty: return df
    out = []
    for (school, email), g in df.groupby(['School', 'Email']):
        first = g.iloc[0].to_dict()
        display_role = collapse_boys_girls_ad_per_person(g['Role'].tolist())
        first['Role'] = display_role
        out.append(first)
    return pd.DataFrame(out)

def apply_table_formatting(file_path):
    wb = load_workbook(file_path)
    used = set()
    for ws in wb.worksheets:
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 2 or max_col < 1:
            continue
        table_range = f"A1:{get_column_letter(max_col)}{max_row}"
        base = re.sub(r"\W+", "", ws.title)[:25]
        name = base + "Tbl"; k = 1
        while name in used:
            name = f"{base}{k}Tbl"; k += 1
        used.add(name)
        if not ws.tables:
            t = Table(displayName=name, ref=table_range)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(t)
        for idx in range(1, max_col+1):
            col = get_column_letter(idx)
            m = max(len(str(c.value)) if c.value else 0 for c in ws[col])
            ws.column_dimensions[col].width = min(max(m + 2, 8), 60)
    wb.save(file_path)

# ---------- WIAA scraping ----------
def extract_school_data(url, school):
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        resp = requests.get(url, headers=headers, timeout=12)
        resp.raise_for_status()
    except requests.RequestException:
        print(f"❌ Could not fetch: {url}")
        return [], []
    soup = BeautifulSoup(resp.text, "html.parser")
    admins, coaches = [], []

    admin_table = soup.find("table", {"id": "tblAdminList"})
    if admin_table:
        for row in admin_table.find_all("tr")[1:]:
            cols = row.find_all("td")
            if len(cols) >= 3:
                role = cols[1].get_text(strip=True).title()
                full_name = cols[2].get_text(" ", strip=True)
                first, last = split_name(full_name)
                email_tag = cols[3].find("span", {"class": "__cf_email__"})
                email = (decode_email(email_tag["data-cfemail"])
                         if email_tag else cols[3].get_text(strip=True))
                admins.append({
                    "School": school.title(),
                    "Role": canonical_admin_role_preserve_gender(role),
                    "First Name": first,
                    "Last Name": last,
                    "Email": email,
                    "State": "WI",
                })

    coach_table = soup.find("table", {"id": "tblCoachList"})
    if coach_table:
        for row in coach_table.find_all("tr")[1:]:
            cols = row.find_all("td")
            if len(cols) >= 4:
                sport = cols[1].get_text(strip=True)
                full_name = cols[2].get_text(" ", strip=True)
                first, last = split_name(full_name)
                role = cols[3].get_text(strip=True).title()
                email_tag = cols[4].find("span", {"class": "__cf_email__"})
                email = (decode_email(email_tag["data-cfemail"])
                         if email_tag else cols[4].get_text(strip=True))
                coaches.append({
                    "School": school.title(),
                    "Sport": sport,
                    "First Name": first,
                    "Last Name": last,
                    "Role": role,
                    "Email": email,
                    "State": "WI",
                })
    return admins, coaches

# ---------- IHSA ----------
def parse_ihsa_jobtitle_to_role_sport(job_title: str):
    if not job_title: return (None, None)
    jt = _norm(job_title); lower = jt.lower()
    gender = None
    if re.search(r"\bboys\b", lower):
        gender = "Boys"; jt = re.sub(r"(?i)\bboys\b", "", jt); lower = jt.lower()
    if re.search(r"\bgirls\b", lower):
        gender = "Girls"; jt = re.sub(r"(?i)\bgirls\b", "", jt); lower = jt.lower()
    if "head coach" in lower:
        sport = re.sub(r"(?i)\bhead\s*coach\b", "", jt); role = "Head Coach"
    elif "coach" in lower and "assistant" not in lower:
        sport = re.sub(r"(?i)\bcoach\b", "", jt); role = "Coach"
    elif ("director" in lower and
          "athletic director" not in lower and
          "activities director" not in lower):
        sport = re.sub(r"(?i)\bdirector\b", "", jt); role = "Director"
    elif "advisor" in lower or "adviser" in lower:
        sport = re.sub(r"(?i)\badvis(e|o)r\b", "", jt); role = "Advisor"
    else:
        return (None, None)
    sport = _title(sport.strip(" -"))
    if gender: sport = f"{gender} {sport}".strip()
    return (role, sport)

def ihsa_rows_to_wiaa_lists(ihsa_df: pd.DataFrame):
    admins, coaches = [], []
    for _, r in ihsa_df.iterrows():
        job   = _norm(r.get("Job Title",""))
        fn    = _title(r.get("First Name",""))
        ln    = _title(r.get("Last Name",""))
        em    = _norm(r.get("Email",""))
        sch   = _title(r.get("School",""))
        state = state_code(r.get("State", "IL"), default="IL")
        if not em:
            continue
        role, sport = parse_ihsa_jobtitle_to_role_sport(job)
        if role and sport:
            coaches.append({
                "School": sch, "Sport": sport,
                "First Name": fn, "Last Name": ln,
                "Role": role, "Email": em, "State": state
            })
        else:
            admins.append({
                "School": sch,
                "Role": canonical_admin_role_preserve_gender(job),
                "First Name": fn, "Last Name": ln,
                "Email": em, "State": state
            })
    return admins, coaches

# ---------- MAIN ----------
def main():
    try:
        xls = pd.ExcelFile(EXCEL_INPUT)
        sheet = SHEET_NAME if SHEET_NAME in xls.sheet_names else xls.sheet_names[0]
        df_schools = pd.read_excel(xls, sheet_name=sheet, dtype=str).fillna("")
    except Exception as e:
        raise RuntimeError(f"Failed to read schools list from {EXCEL_INPUT}: {e}")

    if "Schools" not in df_schools.columns or "School Website" not in df_schools.columns:
        raise RuntimeError("Input sheet must contain columns: 'Schools' and 'School Website'.")

    admin_all, coach_all = [], []

    # WIAA scrape
    for _, row in df_schools.iterrows():
        school = row["Schools"]
        url    = row["School Website"]
        print(f"[WIAA] Scraping: {school}")
        a, c = extract_school_data(url, school)
        admin_all.extend(a)
        coach_all.extend(c)
        time.sleep(1.2)

    # IHSA merge
    if os.path.exists(IHSA_COMBINED_XLSX):
        print(f"[IHSA] Loading combined: {IHSA_COMBINED_XLSX}")
        ihsa_xls = pd.ExcelFile(IHSA_COMBINED_XLSX)
        sh = "All" if "All" in ihsa_xls.sheet_names else ihsa_xls.sheet_names[0]
        ihsa = pd.read_excel(ihsa_xls, sheet_name=sh, dtype=str).fillna("")
        keep = [c for c in ("Job Title","First Name","Last Name","Email","School","State") if c in ihsa.columns]
        ihsa = ihsa[keep].copy()
        ihsa_admins, ihsa_coaches = ihsa_rows_to_wiaa_lists(ihsa)
        print(f"[IHSA] Mapped rows — Admins: {len(ihsa_admins)}, Coaches: {len(ihsa_coaches)}")
        admin_all.extend(ihsa_admins)
        coach_all.extend(ihsa_coaches)
    else:
        print(f"[IHSA] Combined file not found — continuing with WIAA only.")

    # Process admins
    df_admins = pd.DataFrame(admin_all)
    if not df_admins.empty:
        df_admins = deduplicate_admins_by_school_email(df_admins)
        df_athletic_admins = df_admins[df_admins["Role"].isin(ATHLETIC_AD_SHEET_ROLES)].copy()
        df_other_admins    = df_admins[~df_admins["Role"].isin(ATHLETIC_AD_SHEET_ROLES)].copy()
    else:
        df_athletic_admins = pd.DataFrame()
        df_other_admins    = pd.DataFrame()

    # Write workbook
    Path(FINAL_XLSX).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(FINAL_XLSX, engine="openpyxl") as writer:
        if not df_athletic_admins.empty:
            cols = ["School","Role","First Name","Last Name","Email","State"]
            df_athletic_admins = df_athletic_admins.reindex(columns=cols)
            df_athletic_admins.sort_values(['State','School']).to_excel(
                writer, sheet_name="Athletic Admins", index=False)

        if not df_other_admins.empty:
            cols = ["School","Role","First Name","Last Name","Email","State"]
            df_other_admins = df_other_admins.reindex(columns=cols)
            df_other_admins.sort_values(['State','School']).to_excel(
                writer, sheet_name="Administrators", index=False)

        df_coaches = pd.DataFrame(coach_all)
        if not df_coaches.empty:
            df_coaches["Sport Group"] = (
                df_coaches["Sport"]
                .str.replace(r"\b(Boys|Girls)\b", "", regex=True)
                .str.replace(r"[-_]", " ", regex=True)
                .str.strip()
            )
            for sport_group, group in df_coaches.groupby("Sport Group", dropna=False):
                deduped = deduplicate_coaches_by_school_email(group)
                cols = ["School","Sport","First Name","Last Name","Role","Email","State"]
                deduped = deduped.reindex(columns=cols)
                deduped_sorted = deduped.sort_values(['State','School'])
                sheet = re.sub(r"[\\/*?:[\]]", "", (sport_group or "Unknown")).strip().title()[:31] or "Unknown"
                deduped_sorted.to_excel(writer, sheet_name=sheet, index=False)

    apply_table_formatting(FINAL_XLSX)

    # Diff + email + NetSuite sync
    if SEND_EMAIL or SYNC_TO_NETSUITE:
        send  = False
        summary, details = [], []
        all_added_by_sheet   = {}
        all_removed_by_sheet = {}

        def df_to_set(df):
            return set(map(tuple, df.astype(str).to_records(index=False))) if not df.empty else set()

        if os.path.exists(PREV_XLSX):
            prev = pd.read_excel(PREV_XLSX, sheet_name=None)
            curr = pd.read_excel(FINAL_XLSX, sheet_name=None)
            for sheet in curr:
                prev_df = prev.get(sheet, pd.DataFrame())
                curr_df = curr[sheet]
                added   = df_to_set(curr_df) - df_to_set(prev_df)
                removed = df_to_set(prev_df) - df_to_set(curr_df)
                if added or removed:
                    send = True
                    summary.append(f"{sheet} – Added: {len(added)}, Removed: {len(removed)}")
                    all_added_by_sheet[sheet]   = (list(added),   list(curr_df.columns))
                    all_removed_by_sheet[sheet] = (list(removed), list(prev_df.columns if not prev_df.empty else curr_df.columns))
                    for row in added:
                        d = dict(zip(curr_df.columns, row))
                        entry = f"➕ {d.get('First Name','')} {d.get('Last Name','')} – {d.get('Role','')} – {d.get('Email','')} – {d.get('School','')} – {d.get('State','')}"
                        if "Sport" in d: entry += f" – {d.get('Sport','')}"
                        details.append(entry)
                    for row in removed:
                        d = dict(zip(prev_df.columns if not prev_df.empty else curr_df.columns, row))
                        entry = f"➖ {d.get('First Name','')} {d.get('Last Name','')} – {d.get('Role','')} – {d.get('Email','')} – {d.get('School','')} – {d.get('State','')}"
                        if "Sport" in d: entry += f" – {d.get('Sport','')}"
                        details.append(entry)
        else:
            send = True
            summary.append("Initial combined run – no previous version found.")
            # On first run, sync all admin rows to NetSuite
            curr = pd.read_excel(FINAL_XLSX, sheet_name=None)
            for sheet in ["Athletic Admins", "Administrators"]:
                if sheet in curr:
                    curr_df = curr[sheet]
                    all_added_by_sheet[sheet] = (
                        list(df_to_set(curr_df)),
                        list(curr_df.columns)
                    )

        # NetSuite sync
        if SYNC_TO_NETSUITE and (all_added_by_sheet or all_removed_by_sheet):
            for sheet in set(list(all_added_by_sheet.keys()) + list(all_removed_by_sheet.keys())):
                added_rows,   added_cols   = all_added_by_sheet.get(sheet,   ([], []))
                removed_rows, removed_cols = all_removed_by_sheet.get(sheet, ([], []))
                if added_rows or removed_rows:
                    cols = added_cols or removed_cols
                    print(f"\n[NETSUITE] Syncing sheet: {sheet}")
                    sync_changes_to_netsuite(added_rows, removed_rows, cols)

        # Email
        if SEND_EMAIL and send:
            outlook = win32.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To      = EMAIL_TO
            if EMAIL_BCC:
                mail.Bcc = EMAIL_BCC
            mail.Subject = EMAIL_SUBJ
            mail.Body    = (
                "Changes detected in the WI+IL school contacts list:\n\n"
                + "\n".join(summary)
                + "\n\n"
                + "\n".join(details)
            )
            mail.Attachments.Add(os.path.abspath(FINAL_XLSX))
            mail.Send()
            shutil.copy2(FINAL_XLSX, PREV_XLSX)
            print("📬 Email sent with details.")
        elif not send:
            print("✅ No changes detected — no email sent.")
    else:
        print("✉️  Email and sync disabled.")

    print(f"✅ Done. Final combined workbook → {FINAL_XLSX}")

if __name__ == "__main__":
    main()
