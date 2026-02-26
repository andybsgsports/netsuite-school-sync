"""
restore_billto.py
-----------------
Restores Bill-To addresses on all 28 schools.
Reads address from the WIAA scrape data for each school,
adds a Bill-To if one doesn't exist.
"""

import sys, json
sys.path.insert(0, ".")
from netsuite_sync import ns_get, ns_patch, make_auth, BASE_URL, scrape_wiaa_school_detail
import requests
import time
import openpyxl

SPREADSHEET = "School_Master_List.xlsx"


def fetch_with_retry(path, max_retries=3):
    for attempt in range(max_retries):
        try:
            return ns_get(path)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(5 * (attempt + 1))
            else:
                raise


def patch_with_replace(path, body, max_retries=3):
    url = f"{BASE_URL}/{path}?replace=addressBook"
    for attempt in range(max_retries):
        try:
            return requests.patch(url, headers={
                "Authorization": make_auth("PATCH", url),
                "Content-Type": "application/json"
            }, json=body)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(5 * (attempt + 1))
            else:
                raise


def get_schools_from_spreadsheet():
    """Read school data from spreadsheet."""
    wb = openpyxl.load_workbook(SPREADSHEET)
    ws = wb["Schools"]
    headers = [c.value for c in ws[1]]
    schools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        ns_id = data.get("NS Customer ID")
        if not ns_id:
            continue
        schools.append({
            "name": data.get("School Name", ""),
            "ns_id": int(ns_id),
            "url": data.get("School URL", ""),
            "address1": data.get("Address1", ""),
            "address2": data.get("Address2", ""),
            "city": data.get("City", ""),
            "state": data.get("State", "WI"),
            "zip": str(data.get("Zip", "")),
        })
    return schools


def restore_billto(school):
    name = school["name"]
    ns_id = school["ns_id"]
    print(f"\n  {name} ({ns_id}):")

    # Fetch current addresses
    r = fetch_with_retry(f"customer/{ns_id}?expandSubResources=true")
    if r.status_code != 200:
        print(f"    ERROR fetching: {r.status_code}")
        return False

    data = r.json()
    items = data.get("addressBook", {}).get("items", [])

    # Check if Bill-To exists
    has_billing = any(item.get("defaultBilling", False) for item in items)
    if has_billing:
        print(f"    Bill-To already exists — skipping")
        return False

    # Get address from spreadsheet
    addr1 = school.get("address1", "")
    addr2 = school.get("address2", "")
    city = school.get("city", "")
    state = school.get("state", "WI")
    zipcode = school.get("zip", "")

    # If no address in spreadsheet, try scraping
    if not addr1 or not city:
        url = school.get("url", "")
        if url:
            print(f"    No address in spreadsheet, scraping WIAA...")
            info, _, _ = scrape_wiaa_school_detail(url)
            addr1 = info.get("address1", "")
            addr2 = info.get("address2", "")
            city = info.get("city", "")
            state = info.get("state", state)
            zipcode = info.get("zip", "")

    if not addr1 or not city:
        print(f"    No address data available — MANUAL FIX NEEDED")
        return False

    # Build Bill-To (use PO Box/addr2 if available, otherwise addr1)
    bill_addr = addr2 if addr2 else addr1
    bill_to = {
        "defaultShipping": False,
        "defaultBilling": True,
        "label": "Bill-To",
        "addressbookAddress": {
            "addr1": bill_addr,
            "city": city,
            "state": state,
            "zip": zipcode,
            "country": {"id": "US"},
        }
    }

    # Insert Bill-To at the front
    new_items = [bill_to] + items
    print(f"    Adding Bill-To: {bill_addr}, {city}, {state} {zipcode}")

    r = patch_with_replace(f"customer/{ns_id}", {"addressBook": {"items": new_items}})
    if r.status_code == 204:
        print(f"    Done!")
        return True
    else:
        print(f"    ERROR: {r.status_code} {r.text[:300]}")
        return False


def main():
    print("=" * 60)
    print("BILL-TO ADDRESS RESTORATION")
    print("=" * 60)

    schools = get_schools_from_spreadsheet()
    print(f"Found {len(schools)} schools")

    restored = 0
    for school in sorted(schools, key=lambda s: s["name"]):
        if restore_billto(school):
            restored += 1
        time.sleep(2)

    print(f"\nDone. {restored} Bill-To addresses restored.")


if __name__ == "__main__":
    main()
