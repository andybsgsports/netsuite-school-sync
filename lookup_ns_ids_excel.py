"""
lookup_ns_ids_excel.py
----------------------
Searches NetSuite by school name and writes NS Customer IDs to School_Master_List.xlsx.
Run once, then run_sync.py will work.

  python lookup_ns_ids_excel.py
"""

import openpyxl
import requests
import time
from netsuite_sync import make_auth, NS_ACCOUNT

EXCEL_FILE = "School_Master_List.xlsx"
BASE_URL   = f"https://{NS_ACCOUNT}.suitetalk.api.netsuite.com/services/rest/record/v1"

def ns_search(school_name):
    encoded = requests.utils.quote(school_name)
    url = f"{BASE_URL}/customer?q=companyName+CONTAINS+{encoded}&limit=10&isInactive=false"
    resp = requests.get(url, headers={"Authorization": make_auth("GET", url), "Content-Type": "application/json"})
    if resp.status_code == 200:
        return resp.json().get("items", [])
    print(f"  ERROR {resp.status_code}: {resp.text[:200]}")
    return []

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb["Schools"]
rows = list(ws.iter_rows(values_only=True))
hdr  = list(rows[0])

name_col = hdr.index("School Name")
id_col   = hdr.index("NS Customer ID")

print("Looking up NetSuite Customer IDs...\n")
updated = 0

for row_idx, row in enumerate(rows[1:], start=2):
    school_name = (row[name_col] or "").strip()
    current_id  = str(row[id_col] or "").strip()
    if not school_name:
        continue
    if current_id:
        print(f"  [SKIP]  {school_name} already has ID: {current_id}")
        continue

    results = ns_search(school_name)
    if len(results) == 1:
        ns_id = str(results[0]["id"])
        print(f"  [FOUND] {school_name} -> {ns_id} ({results[0].get('companyName','')})")
        ws.cell(row=row_idx, column=id_col + 1, value=ns_id)
        updated += 1
    elif len(results) > 1:
        print(f"  [MULTI] {school_name} -> {len(results)} matches:")
        for r in results:
            print(f"           {r['id']}  {r.get('companyName','')}")
    else:
        print(f"  [MISS]  {school_name}")
    time.sleep(0.3)

wb.save(EXCEL_FILE)
print(f"\nDone! {updated} IDs written to {EXCEL_FILE}")
