"""
rep_digests.py
--------------
Consolidated replacement for the six per-rep WIAA scraper scripts that used to
live on the Desktop and send Outlook emails via Task Scheduler.

One run:
  1. Reads the main master sheet's Schools tab, grouped by Sales Rep col
  2. For each configured rep, scrapes WIAA for their schools
  3. Builds an xlsx with Athletic Admins / Administrators / per-sport coach tabs
  4. Diffs current vs. previous snapshot (stored in snapshots/{rep}.json)
  5. Emails the rep (and BCC andy) via Gmail SMTP if anything changed
  6. Writes the new snapshot back to disk (committed by the workflow)
  7. Uploads the per-rep xlsx to the shared Drive archive folder
  8. Merges any new scraped rows into the master sheet's Contacts tab

Env vars required:
  GOOGLE_SHEET_ID          - main master sheet (Schools tab + Contacts tab)
  GOOGLE_CREDENTIALS_JSON  - service account JSON (same as daily-sync)
  GMAIL_USER               - e.g. andy@bsgsports.com
  GMAIL_APP_PASSWORD       - Gmail app password (requires 2FA enabled)
  DRY_RUN                  - if "1", send all emails to GMAIL_USER only
  REP_FILTER               - if set, only process that rep name (testing)
"""

import io
import json
import os
import re
import smtplib
import sys
import time
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

import gspread
import pandas as pd
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from netsuite_sync import scrape_wiaa_school_detail
from ihsa_sync import fetch_school_staff, fetch_email, extract_school_id

# -- Config -------------------------------------------------------------------
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
DELAY_BETWEEN_SCHOOLS = 1.2  # seconds

SNAPSHOT_DIR = Path(__file__).parent / "snapshots"

# Roles that go on the Athletic Admins tab (vs. generic Administrators).
ATHLETIC_AD_ROLES = {
    "Athletic Director",
    "Assistant Principal, Athletic Director",
    "Boys Athletic Director",
    "Girls Athletic Director",
}

# Rep-to-email mapping. "name" must match the value in the sheet's Sales Rep
# column exactly. "cc" is optional.
#
# TODO(andy): confirm the addresses for Howie, JohnV, Tyler, Wedge. These are
# derived from the old per-rep scripts where available and my best guess
# otherwise. Update before flipping DRY_RUN off.
REPS = [
    # Andy also gets IL schools (IHSA API) in his digest — only rep with IL.
    {"name": "Andrew Murray", "email": "andy@bsgsports.com",   "cc": None, "include_il": True},
    {"name": "Jeff Howard",   "email": "howie@bsgsports.com",  "cc": None},
    {"name": "Tyler Fuhrman", "email": "tyler@bsgsports.com",  "cc": None},
    {"name": "Kyle Loughrin", "email": "kylel@bsgsports.com",  "cc": None},
    {"name": "Paul Speth",    "email": "paul@bsgsports.com",   "cc": "julie@bsgsports.com"},
    {"name": "John Viles",    "email": "johnv@bsgsports.com",  "cc": None},
    {"name": "Jeff Wedvick",  "email": "wedge@bsgsports.com",  "cc": None},
]

GOOGLE_SHEET_ID_MAIN = os.environ.get("GOOGLE_SHEET_ID", "")  # for IL_Schools tab

# Shared Drive folder where each rep's XLSX is archived after emailing.
# Default = Andy's "School/Contact Sync - WIAA/IHSA Scrape - BSG to Netsuite"
# folder. Override with GOOGLE_DRIVE_DIGEST_FOLDER_ID env var.
DRIVE_DIGEST_FOLDER_ID = os.environ.get(
    "GOOGLE_DRIVE_DIGEST_FOLDER_ID",
    "1ZcchdQmDngJc_sro-LyK8yAYofdwkLTK",
).strip()

DRY_RUN = os.environ.get("DRY_RUN", "") == "1"
REP_FILTER = os.environ.get("REP_FILTER", "").strip()


# -- Google Sheets -----------------------------------------------------------
def get_gspread_client():
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
    if creds_json:
        creds = Credentials.from_service_account_info(
            json.loads(creds_json), scopes=GOOGLE_SCOPES
        )
    else:
        creds_file = Path(__file__).parent / "credentials.json"
        creds = Credentials.from_service_account_file(str(creds_file), scopes=GOOGLE_SCOPES)
    return gspread.authorize(creds)


def load_rep_schools(gc):
    """Returns {rep_name: [(school_name, school_url), ...]} for WI rows on
    the MAIN master sheet's Schools tab. The canonical "School Name"
    column is used so scraped contact rows land with the same name the
    rest of the system uses (e.g. 'Antioch Community High School', not
    the legacy shorthand 'Antioch' from the retired WI School List-
    Master sheet that this function used to read)."""
    if not GOOGLE_SHEET_ID_MAIN:
        return {}
    wb = gc.open_by_key(GOOGLE_SHEET_ID_MAIN)
    try:
        ws = wb.worksheet("Schools")
    except Exception:
        return {}
    records = ws.get_all_records()
    # Name collisions (one School Name -> two NS customers, e.g. the two
    # Pecatonicas) blend two schools' rosters — skip those names entirely.
    from school_netsuite_sync import screen_school_name_collisions
    quarantined = screen_school_name_collisions(
        [(r.get("School Name", ""), r.get("NS Customer ID", ""))
         for r in records], log_prefix="[digests][schools]")
    by_rep = {}
    seen = set()
    for row in records:
        state = str(row.get("State", "")).strip().upper()
        if state != "WI":
            continue
        school = str(row.get("School Name", "")).strip()
        url    = str(row.get("School URL", "")).strip()
        rep    = str(row.get("Sales Rep", "")).strip()
        locked = str(row.get("Locked", "")).strip().upper() == "Y"
        if not (school and url and rep) or locked:
            continue
        if school in quarantined or (school, url) in seen:
            continue
        seen.add((school, url))
        by_rep.setdefault(rep, []).append((school, url))
    return by_rep


def load_il_schools(gc):
    """Returns [(school_name, school_website), ...] — IL rows from the
    unified Schools tab on the main master sheet (State == 'IL').
    Previously read from a separate IL_Schools tab; that tab was retired
    when WI + IL consolidated into one Schools tab."""
    if not GOOGLE_SHEET_ID_MAIN:
        return []
    wb = gc.open_by_key(GOOGLE_SHEET_ID_MAIN)
    try:
        ws = wb.worksheet("Schools")
    except Exception:
        return []
    records = ws.get_all_records()
    from school_netsuite_sync import screen_school_name_collisions
    quarantined = screen_school_name_collisions(
        [(r.get("School Name", ""), r.get("NS Customer ID", ""))
         for r in records], log_prefix="[digests][schools]")
    out = []
    seen = set()
    for row in records:
        state = str(row.get("State", "")).strip().upper()
        if state != "IL":
            continue
        school = str(row.get("School Name", "")).strip()
        url    = str(row.get("School URL", "")).strip()
        locked = str(row.get("Locked", "")).strip().upper() == "Y"
        if school and url and not locked \
                and school not in quarantined and (school, url) not in seen:
            seen.add((school, url))
            out.append((school, url))
    return out


# -- Scraping helpers --------------------------------------------------------
def scrape_rep(rep_name, schools):
    """Scrape every school assigned to `rep_name`.
    Returns (admins, coaches, scraped_schools) — scraped_schools is the set of
    (smart-titled) school names that scraped successfully this run. A school
    whose WIAA page errored is NOT in this set, so callers can avoid treating
    its previously-known staff as "departed" just because of a fetch failure."""
    admins, coaches = [], []
    scraped_schools = set()
    for i, (school, url) in enumerate(schools, 1):
        print(f"  [{i}/{len(schools)}] {school}")
        try:
            _info, scraped_admins, scraped_coaches = scrape_wiaa_school_detail(url)
        except Exception as exc:
            print(f"    ERROR: {exc}")
            continue
        scraped_schools.add(smart_title(school))
        for a in scraped_admins:
            admins.append({
                "School":     smart_title(school),
                "Role":       canonical_admin_role(a.get("role", "")),
                "First Name": smart_title(a.get("first") or ""),
                "Last Name":  smart_title(a.get("last") or ""),
                "Email":      a.get("email", ""),
                "State":      "WI",
            })
        for c in scraped_coaches:
            coaches.append({
                "School":     smart_title(school),
                "Sport":      c.get("role", ""),  # netsuite_sync returns sport in role
                "First Name": smart_title(c.get("first") or ""),
                "Last Name":  smart_title(c.get("last") or ""),
                "Role":       c.get("type", ""),  # Head Coach / Assistant Coach / Coach
                "Email":      c.get("email", ""),
                "State":      "WI",
            })
        time.sleep(DELAY_BETWEEN_SCHOOLS)
    return dedup_admins(admins), dedup_coaches(coaches), scraped_schools


# IHSA role IDs that belong on the Administrators-style sheets rather than a
# sport-coach sheet. Prefix meanings: A* / B* = Admin, G* = Medical, everything
# else is a coach / activity head.
IL_ADMIN_PREFIXES = ("A", "B", "G")
# Admin role-IDs that specifically belong on the Athletic Admins sheet
IL_ATHLETIC_AD_ROLE_IDS = {"B2-AthDir", "C1-BoysAD", "C1-GirlsAD"}


def scrape_il_schools(il_schools):
    """
    Scrape IL via IHSA API. Returns (admins, coaches, scraped_schools) in the
    same shape as scrape_rep() so they can be merged into Andy's combined
    xlsx and the same departure-scoping logic applies to IL schools too.
    """
    admins, coaches = [], []
    scraped_schools = set()
    for i, (school, url) in enumerate(il_schools, 1):
        school_id = extract_school_id(url)
        if not school_id:
            print(f"  [IL {i}/{len(il_schools)}] {school}  -- can't parse id, skip")
            continue
        print(f"  [IL {i}/{len(il_schools)}] {school} (id {school_id})")
        try:
            people = fetch_school_staff(school_id)
        except Exception as exc:
            print(f"    ERROR staff2: {exc}")
            continue
        scraped_schools.add(smart_title(school))
        # Resolve emails
        for p in people:
            if p.get("has_email") and p.get("person_id"):
                try:
                    p["email"] = fetch_email(school_id, p["person_id"])
                except Exception:
                    p["email"] = ""
                time.sleep(0.15)
        for p in people:
            if not p.get("email"):
                continue
            role_id = p.get("role_id", "") or ""
            role_name = (p.get("role") or "").strip()
            coach_type = p.get("type", "")
            if not role_name:
                continue
            first = smart_title(p.get("first") or "")
            last  = smart_title(p.get("last") or "")
            if coach_type == "Admin":
                admins.append({
                    "School":     smart_title(school),
                    "Role":       canonical_admin_role(role_name) if role_id in IL_ATHLETIC_AD_ROLE_IDS else smart_title(role_name),
                    "First Name": first,
                    "Last Name":  last,
                    "Email":      p["email"],
                    "State":      "IL",
                })
            else:
                # role_name is the sport already (e.g. "Boys Baseball")
                coaches.append({
                    "School":     smart_title(school),
                    "Sport":      smart_title(role_name),
                    "First Name": first,
                    "Last Name":  last,
                    "Role":       coach_type,   # "Head Coach" / "Assistant Coach" / "Coach"
                    "Email":      p["email"],
                    "State":      "IL",
                })
        time.sleep(0.5)
    return dedup_admins(admins), dedup_coaches(coaches), scraped_schools


def _norm(s):
    return re.sub(r"\s+", " ", ("" if s is None else str(s)).strip())


def smart_title(s):
    """
    Like str.title() but:
      - doesn't capitalize after apostrophes ("Principal's", not "Principal'S")
      - leaves already-mixed-case input alone ("McDonald" stays "McDonald")
    """
    t = str(s or "")
    if not t:
        return t
    # If the string has any mixed case (e.g. "McDonald", "D'Andrea"), preserve it.
    if t != t.lower() and t != t.upper():
        return t
    return re.sub(r"\b[a-zA-Z]+(?:'[a-zA-Z]+)?",
                  lambda m: m.group(0)[0].upper() + m.group(0)[1:].lower(),
                  t)


def canonical_admin_role(role):
    r = _norm(role)
    low = r.lower()
    if "assistant principal" in low and "athletic director" in low:
        return "Assistant Principal, Athletic Director"
    if "assistant athletic director" in low:
        return "Assistant Athletic Director"
    if "activities director" in low:
        return "Activities Director"
    if "supervisor" in low:
        return smart_title(r)
    if "athletic director" in low and "assistant" not in low:
        if "boys" in low:
            return "Boys Athletic Director"
        if "girls" in low:
            return "Girls Athletic Director"
        return "Athletic Director"
    return smart_title(r)


def dedup_admins(admins):
    """Collapse per (School, Email). Combine gendered ADs into plain AD."""
    if not admins:
        return []
    df = pd.DataFrame(admins)
    out = []
    for (_school, _email), group in df.groupby(["School", "Email"]):
        row = group.iloc[0].to_dict()
        roles = set(group["Role"].tolist())
        if {"Boys Athletic Director", "Girls Athletic Director"} <= roles or "Athletic Director" in roles:
            row["Role"] = "Athletic Director"
        elif len(roles) > 1:
            row["Role"] = " & ".join(sorted(roles))
        out.append(row)
    return out


def sport_group_of(sport):
    """Canonical sport group: strip Boys/Girls, separators, collapse whitespace."""
    s = re.sub(r"\b(Boys|Girls)\b", "", str(sport), flags=re.IGNORECASE)
    s = re.sub(r"[-_&]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return smart_title(s)


def dedup_coaches(coaches):
    """
    Dedup per (School, Email, SportGroup) so one coach covering Boys+Girls
    of the same sport becomes a single row, but a coach covering different
    sports (e.g. Basketball + Golf) stays as separate rows — one per sheet.
    """
    if not coaches:
        return []
    df = pd.DataFrame(coaches)
    df["SportGroup"] = df["Sport"].map(sport_group_of)
    out = []
    for (_school, _email, sg), group in df.groupby(["School", "Email", "SportGroup"]):
        row = group.iloc[0].to_dict()
        row["SportGroup"] = sg
        roles = {str(r) for r in group["Role"].tolist()}
        sports = list(dict.fromkeys(group["Sport"].tolist()))
        if "Head Coach" in roles:
            row["Role"] = "Head Coach"
        elif "Assistant Coach" in roles:
            row["Role"] = "Assistant Coach"
        elif "Coach" in roles:
            row["Role"] = "Coach"
        has_boys = any("boys" in s.lower() for s in sports)
        has_girls = any("girls" in s.lower() for s in sports)
        if has_boys and has_girls:
            row["Sport"] = f"Boys & Girls {sg}".strip()
        elif len(sports) == 1:
            row["Sport"] = sports[0]
        else:
            # Same SportGroup, multiple variants (rare) — pick the cleanest.
            row["Sport"] = sports[0]
        out.append(row)
    return out


# -- XLSX output -------------------------------------------------------------
def build_xlsx(admins, coaches, rep_name):
    """Build the per-rep xlsx in memory. Returns (bytes, sheet_summary_dict)."""
    bio = io.BytesIO()
    summary = {}
    df_admins = pd.DataFrame(admins)
    df_coaches = pd.DataFrame(coaches)

    wrote_any = False
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        if not df_admins.empty:
            df_ath = df_admins[df_admins["Role"].isin(ATHLETIC_AD_ROLES)]
            df_oth = df_admins[~df_admins["Role"].isin(ATHLETIC_AD_ROLES)]
            cols = ["School", "Role", "First Name", "Last Name", "Email", "State"]
            if not df_ath.empty:
                df_ath.reindex(columns=cols).sort_values(["State", "School"]) \
                    .to_excel(w, sheet_name="Athletic Admins", index=False)
                summary["Athletic Admins"] = len(df_ath)
                wrote_any = True
            if not df_oth.empty:
                df_oth.reindex(columns=cols).sort_values(["State", "School"]) \
                    .to_excel(w, sheet_name="Administrators", index=False)
                summary["Administrators"] = len(df_oth)
                wrote_any = True

        if not df_coaches.empty:
            df_coaches = df_coaches.copy()
            if "SportGroup" not in df_coaches.columns:
                df_coaches["SportGroup"] = df_coaches["Sport"].map(sport_group_of)
            cols = ["School", "Sport", "First Name", "Last Name", "Role", "Email", "State"]
            for sport_group, group in df_coaches.groupby("SportGroup", dropna=False):
                sheet = re.sub(r"[\\/*?:[\]]", "", sport_group or "Unknown").strip()[:31] or "Unknown"
                df_group = group.reindex(columns=cols).sort_values(["State", "School"])
                df_group.to_excel(w, sheet_name=sheet, index=False)
                summary[sheet] = len(df_group)
                wrote_any = True

        # openpyxl refuses to save a workbook with zero sheets
        # ("At least one sheet must be visible"). When a scrape comes back
        # completely empty (e.g. WIAA returning 503s for every school), write
        # a placeholder sheet so the run doesn't crash and take down every
        # rep processed after this one.
        if not wrote_any:
            pd.DataFrame({"Note": ["No contacts scraped this run (source site "
                                   "may have been unavailable)."]}) \
                .to_excel(w, sheet_name="No Data", index=False)

    apply_table_formatting(bio)
    return bio.getvalue(), summary


def apply_table_formatting(bio):
    bio.seek(0)
    wb = load_workbook(bio)
    used = set()
    for ws in wb.worksheets:
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 2 or max_col < 1:
            continue
        base = re.sub(r"\W+", "", ws.title)[:25] or "Data"
        name = base + "Tbl"
        k = 1
        while name in used:
            name = f"{base}{k}Tbl"
            k += 1
        used.add(name)
        if not ws.tables:
            t = Table(displayName=name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(t)
        for idx in range(1, max_col + 1):
            col = get_column_letter(idx)
            width = max((len(str(c.value)) if c.value else 0) for c in ws[col])
            ws.column_dimensions[col].width = min(max(width + 2, 8), 60)
    bio.seek(0)
    bio.truncate()
    wb.save(bio)


# -- Snapshots + diff --------------------------------------------------------
def snapshot_path(rep_name):
    safe = re.sub(r"[^A-Za-z0-9]+", "_", rep_name).strip("_")
    return SNAPSHOT_DIR / f"{safe}.json"


def _diff_key(school, email, role, sport):
    """Case-fold identity parts that aren't guaranteed stable across runs.
    Prevents phantom add/remove churn when the scraper re-cases a role
    ('Ad Admin Assistant' -> 'AD Admin Assistant') or sport ('Track And
    Field' -> 'Track and Field'). School is stable from the Schools tab."""
    return (
        (school or "").strip(),
        (email or "").strip().lower(),
        (role or "").strip().lower(),
        (sport or "").strip().lower(),
    )


def contacts_to_records(admins, coaches):
    """
    Returns {case-folded key: {"first":, "last":, "email":, "role":, "sport":}}.
    Key is the stable identity used for diffing; value holds display-cased
    fields for the email lines (so a re-case doesn't cause phantom churn).
    """
    recs = {}
    for a in admins:
        key = _diff_key(a["School"], a["Email"], a["Role"], "")
        recs[key] = {"first": a.get("First Name", ""), "last": a.get("Last Name", ""),
                     "email": a.get("Email", ""), "role": a.get("Role", ""), "sport": ""}
    for c in coaches:
        key = _diff_key(c["School"], c["Email"], c["Role"], c.get("Sport", ""))
        recs[key] = {"first": c.get("First Name", ""), "last": c.get("Last Name", ""),
                     "email": c.get("Email", ""), "role": c.get("Role", ""),
                     "sport": c.get("Sport", "")}
    return recs


def load_snapshot(rep_name):
    """Returns (keyset, records_dict) or (None, {})."""
    p = snapshot_path(rep_name)
    if not p.exists():
        return None, {}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        # New format: records is a list of {school, email, role, sport, first, last}
        if "records" in data:
            recs = {}
            for r in data["records"]:
                key = _diff_key(r.get("school", ""), r.get("email", ""),
                                r.get("role", ""), r.get("sport", ""))
                recs[key] = {"first": r.get("first", ""), "last": r.get("last", ""),
                             "email": r.get("email", ""), "role": r.get("role", ""),
                             "sport": r.get("sport", "")}
            return set(recs.keys()), recs
        # Legacy format: keys-only list of tuples, no name info
        return {_diff_key(*k) for k in data.get("keys", []) if len(k) == 4}, {}
    except Exception:
        return None, {}


def save_snapshot(rep_name, records):
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    p = snapshot_path(rep_name)
    # Store display-cased values; _diff_key re-normalizes on read so case
    # in the JSON doesn't affect matching.
    serializable = [
        {"school": k[0],
         "email":  v.get("email", k[1]),
         "role":   v.get("role", k[2]),
         "sport":  v.get("sport", k[3]),
         "first":  v.get("first", ""),
         "last":   v.get("last", "")}
        for k, v in sorted(records.items())
    ]
    p.write_text(
        json.dumps(
            {
                "rep": rep_name,
                "updated": datetime.utcnow().isoformat() + "Z",
                "records": serializable,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def diff_keys(previous, current):
    if previous is None:
        return set(), set(), True  # first run
    added = current - previous
    removed = previous - current
    return added, removed, False


# -- Email -------------------------------------------------------------------
def _drive_service():
    """Service-account-authenticated Google Drive v3 client."""
    from googleapiclient.discovery import build
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
    if creds_json:
        creds = Credentials.from_service_account_info(
            json.loads(creds_json), scopes=GOOGLE_SCOPES)
    else:
        creds = Credentials.from_service_account_file(
            str(Path(__file__).parent / "credentials.json"),
            scopes=GOOGLE_SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _find_or_create_subfolder(drive, name, parent_id):
    """Get id of `name` folder inside `parent_id`, creating it if absent."""
    safe = name.replace("'", "\\'")
    q = (f"'{parent_id}' in parents and "
         f"mimeType='application/vnd.google-apps.folder' "
         f"and name='{safe}' and trashed=false")
    res = drive.files().list(
        q=q, fields="files(id)", supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    if res.get("files"):
        return res["files"][0]["id"]
    meta = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    out = drive.files().create(body=meta, fields="id",
                               supportsAllDrives=True).execute()
    return out["id"]


def upload_digest_to_drive(rep_name, xlsx_bytes, xlsx_name):
    """Archive the rep's XLSX in a per-rep subfolder of the shared Drive
    folder. No-op if DRIVE_DIGEST_FOLDER_ID isn't set or the folder isn't
    shared with the service account."""
    from googleapiclient.http import MediaIoBaseUpload
    if not DRIVE_DIGEST_FOLDER_ID:
        return
    try:
        drive = _drive_service()
        rep_folder = _find_or_create_subfolder(drive, rep_name, DRIVE_DIGEST_FOLDER_ID)
        dated_name = f"{datetime.now().strftime('%Y-%m-%d')}__{xlsx_name}"
        media = MediaIoBaseUpload(
            io.BytesIO(xlsx_bytes),
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet"),
            resumable=False,
        )
        drive.files().create(
            body={"name": dated_name, "parents": [rep_folder]},
            media_body=media,
            fields="id",
            supportsAllDrives=True,
        ).execute()
        print(f"  [Drive] Archived {dated_name} -> {rep_name}/")
    except Exception as e:
        print(f"  [Drive] WARN upload failed for {rep_name}: {e}")


def send_email(rep, subject, body, xlsx_bytes, xlsx_name):
    """
    Recipient logic:
      DIGESTS_OVERRIDE_TO set  -> all emails go there, subject gets [NEW SYS] tag
                                  (shadow mode for parallel validation)
      DRY_RUN=1                -> all emails go to GMAIL_USER, labeled [DRY RUN]
      otherwise (true live)    -> rep's actual email + CC
    """
    gmail_user = os.environ.get("GMAIL_USER", "")
    gmail_pw = os.environ.get("GMAIL_APP_PASSWORD", "")
    override_to = os.environ.get("DIGESTS_OVERRIDE_TO", "").strip()
    if not (gmail_user and gmail_pw):
        print("  WARNING: GMAIL_USER / GMAIL_APP_PASSWORD not set -- skipping send")
        return False

    if override_to:
        to_addr = override_to
        cc_addr = None
        bcc_addr = None  # primary recipient is already the user — skip BCC
        subject = f"[NEW SYS] {subject}"
        body = (
            f"(Shadow-mode email from GitHub Actions; would have gone to {rep['email']}"
            + (f", CC {rep['cc']}" if rep.get("cc") else "")
            + ".)\n\n" + body
        )
    elif DRY_RUN:
        to_addr = gmail_user
        cc_addr = None
        bcc_addr = None  # TO is already GMAIL_USER — skip BCC
        body = (
            f"[DRY RUN — would send to {rep['email']}"
            + (f", CC {rep['cc']}" if rep.get("cc") else "")
            + "]\n\n" + body
        )
    else:
        to_addr = rep["email"]
        cc_addr = rep.get("cc")
        bcc_addr = gmail_user  # true live: BCC andy so he sees each rep's email

    msg = EmailMessage()
    msg["From"] = gmail_user
    msg["To"] = to_addr
    if cc_addr:
        msg["Cc"] = cc_addr
    msg["Subject"] = subject
    msg.set_content(body)
    msg.add_attachment(
        xlsx_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=xlsx_name,
    )

    recipients = [to_addr]
    if cc_addr:
        recipients.append(cc_addr)
    if bcc_addr and bcc_addr not in recipients:
        recipients.append(bcc_addr)

    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        s.starttls()
        s.login(gmail_user, gmail_pw)
        s.send_message(msg, to_addrs=recipients)
    print(f"  Email sent to {to_addr}" + (f" (CC {cc_addr})" if cc_addr else ""))
    return True


# -- Main --------------------------------------------------------------------
# Scraper-owned Type values — the reconcile only ever flips rows the scraper
# itself created. A manually added row with any other Type is never touched.
RECONCILE_TYPES = {"admin", "head coach", "coach"}
# A school is only reconciled when its scrape returned at least this many
# people. A thinner result smells like a partial/failed page, not turnover.
RECONCILE_MIN_PEOPLE = 5


def reconcile_absent_contacts(contacts_data, scraped, scraped_schools):
    """State-based departure sweep, run after the snapshot-delta flip.

    The snapshot diff only reacts to TRANSITIONS — someone present last run
    and gone this run. Anyone who left before the snapshots started
    watching (Steve Gertz was already off Dundee-Crown's IHSA page when the
    departure feature first ran, so no 'removed' delta ever fired for him)
    or whose sheet rows carry a role format the scraper no longer emits
    ('Boys Athletic Director' vs today's 'Athletic Director') can never
    match a delta and sits at Sync=Y forever — NetSuite keeps them active
    with their Ship-To address on the school. This pass closes that hole by
    comparing the sheet against the CURRENT scrape: any Sync=Y row at a
    successfully-scraped school whose email is absent from that school's
    fresh staff list flips to Sync=N. push_only then does the rest exactly
    as for any departure: inactivate (or detach, for co-op people still
    active elsewhere) AND remove their Ship-To address line.

    Safety rails:
      - whole-person: keyed on email, so all of a person's rows at a school
        flip together (push_only's per-email dedupe assumes a person is
        never half-active at one school);
      - only schools in scraped_schools (successful scrape this run) that
        returned >= RECONCILE_MIN_PEOPLE people;
      - only scraper-owned Types (Admin / Head Coach / Coach) — manual rows
        with any other Type are never flipped;
      - reversible: flipping a row back to Y re-syncs the contact with
        isInactive=False on the next push.
    """
    from school_netsuite_sync import C_SCHOOL, C_FIRST, C_LAST, C_EMAIL, C_TYPE, C_SYNC
    if not scraped_schools:
        return 0
    cur = {}
    for _state, rec in scraped:
        s = str(rec.get("School", "")).strip()
        e = str(rec.get("Email", "")).strip().lower()
        if s and e:
            cur.setdefault(s, set()).add(e)
    eligible = {s for s in scraped_schools
                if len(cur.get(s, ())) >= RECONCILE_MIN_PEOPLE}
    thin = set(scraped_schools) - eligible
    if thin:
        print(f"[merge][reconcile] skipping {len(thin)} scraped school(s) with "
              f"<{RECONCILE_MIN_PEOPLE} scraped people (partial-scrape guard): "
              f"{sorted(thin)[:5]}" + (" ..." if len(thin) > 5 else ""))
    flipped = 0
    for c in contacts_data:
        if str(c.get(C_SYNC, "N")).strip().upper() != "Y":
            continue
        if str(c.get(C_TYPE, "")).strip().lower() not in RECONCILE_TYPES:
            continue
        sch = str(c.get(C_SCHOOL, "")).strip()
        em = str(c.get(C_EMAIL, "")).strip().lower()
        if not em or sch not in eligible:
            continue
        if em not in cur[sch]:
            c[C_SYNC] = "N"
            flipped += 1
            print(f"[merge][reconcile] {c.get(C_FIRST, '')} {c.get(C_LAST, '')} "
                  f"<{em}> not on {sch}'s current scrape — Sync=N")
    return flipped


def merge_scraped_into_master_sheet(gc, scraped, departed_triples=None,
                                    scraped_schools=None):
    """Merge today's scraped admins+coaches into the main master sheet's
    Contacts tab. Dedupes on (School, Email, Role) so re-runs don't create
    duplicate rows.

    departed_triples: optional set of (school, email_lower, role_col_lower)
    keys — people who were on last run's snapshot for that school but are
    no longer on this run's fresh WIAA/IHSA scrape (and whose school DID
    scrape successfully this run, so we trust the absence). Any Sync=Y
    Contacts-tab row matching one of these triples is flipped to Sync=N.
    Does NOT touch NS directly; a later workflow (push_only.py) reads the
    Sync column and inactivates the NetSuite contact + Ship-To."""
    if not scraped and not departed_triples:
        return
    main_sheet_id = os.environ.get("GOOGLE_SHEET_ID", "")
    if not main_sheet_id:
        print("\n[merge] GOOGLE_SHEET_ID not set — skipping master sheet update")
        return

    # Import sheet helpers lazily (avoid circular imports at module load)
    from school_netsuite_sync import (
        load_contacts, save_contacts,
        canonicalize_contact_school_names, MASTER_TAB,
        C_SCHOOL, C_FIRST, C_LAST, C_EMAIL, C_ROLE, C_TYPE,
        C_SYNC, C_NS_CID, C_NS_CUS, C_SYNCED,
    )
    wb = gc.open_by_key(main_sheet_id)
    contacts_ws = wb.worksheet("Contacts")
    contacts_data = contacts_ws.get_all_records()

    # Heal school renames BEFORE keying on School Name: scraped records
    # carry the Schools tab's CURRENT name, so rows stranded under a
    # school's old name would never match and every person would be
    # re-added as a duplicate row (and later a duplicate NS contact).
    canonicalize_contact_school_names(
        contacts_data, wb.worksheet(MASTER_TAB).get_all_records(),
        log_prefix="[merge][rename-heal]")

    existing_keys = {
        (str(c.get(C_SCHOOL, "")).strip(),
         str(c.get(C_EMAIL, "")).strip().lower(),
         str(c.get(C_ROLE, "")).strip().lower())
        for c in contacts_data
        if str(c.get(C_EMAIL, "")).strip()
    }

    added = 0
    for state, rec in scraped:
        school = str(rec.get("School", "")).strip()
        email  = str(rec.get("Email", "")).strip()
        # Admins store role in 'Role'; coaches store sport in 'Sport' and
        # "Head Coach"/"Coach" in 'Role'. Normalize to the Contacts-tab
        # convention: C_ROLE holds the sport for coaches, admin title for
        # admins; C_TYPE holds "Head Coach"/"Coach"/"Admin".
        sport = str(rec.get("Sport", "")).strip()
        if sport:
            role_col = sport
            type_col = str(rec.get("Role", "")).strip() or "Coach"
        else:
            role_col = str(rec.get("Role", "")).strip()
            type_col = "Admin"
        if not (school and email and role_col):
            continue
        key = (school, email.lower(), role_col.lower())
        if key in existing_keys:
            continue
        contacts_data.append({
            C_SCHOOL: school,
            C_FIRST:  str(rec.get("First Name", "")).strip(),
            C_LAST:   str(rec.get("Last Name", "")).strip(),
            C_EMAIL:  email,
            C_ROLE:   role_col,
            C_TYPE:   type_col,
            C_SYNC:   "Y",
            C_NS_CID: "",
            C_NS_CUS: "",
            C_SYNCED: "",
        })
        existing_keys.add(key)
        added += 1

    departed = 0
    if departed_triples:
        for c in contacts_data:
            if str(c.get(C_SYNC, "N")).strip().upper() != "Y":
                continue
            triple = (str(c.get(C_SCHOOL, "")).strip(),
                      str(c.get(C_EMAIL, "")).strip().lower(),
                      str(c.get(C_ROLE, "")).strip().lower())
            if triple in departed_triples:
                c[C_SYNC] = "N"
                departed += 1

    reconciled = reconcile_absent_contacts(contacts_data, scraped, scraped_schools)

    save_contacts(contacts_ws, contacts_data)
    print(f"\n[merge] {added} new row(s) added to master sheet Contacts tab")
    if departed_triples:
        print(f"[merge] {departed} row(s) flipped to Sync=N (no longer on WIAA/IHSA) "
              f"— push_only.py will inactivate them in NetSuite tonight")
    if reconciled:
        print(f"[merge] {reconciled} additional stale row(s) reconciled to Sync=N "
              f"(absent from the current scrape) — NS inactivation + Ship-To "
              f"removal happen on tonight's push")


def main():
    print("=" * 60)
    print(f"  Rep Digests  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  DRY_RUN={DRY_RUN}")
    print("=" * 60)

    gc = get_gspread_client()
    by_rep = load_rep_schools(gc)
    print(f"\nReps in sheet: {sorted(by_rep.keys())}")

    # Pre-load IL schools once (used by any rep with include_il=True)
    il_schools = []
    if any(r.get("include_il") for r in REPS):
        il_schools = load_il_schools(gc)
        print(f"IL schools available: {len(il_schools)}")

    rep_name_to_config = {r["name"]: r for r in REPS}
    results = []
    all_scraped = []          # accumulate across reps for one sheet write at the end
    all_departed_triples = set()  # (school, email_lower, role_col_lower) no longer on WIAA/IHSA
    all_scraped_schools = set()   # schools that scraped OK this run (guards the reconcile)

    for rep in REPS:
        if REP_FILTER and rep["name"] != REP_FILTER:
            continue
        schools = by_rep.get(rep["name"], [])
        if not schools:
            print(f"\n[{rep['name']}] No schools in sheet — skipping")
            continue

        print(f"\n{'-' * 60}")
        print(f"[{rep['name']}] {len(schools)} schools")
        print("-" * 60)

        admins, coaches, wi_scraped_schools = scrape_rep(rep["name"], schools)
        # A rep with WI schools that yields ZERO admins AND ZERO coaches is a
        # scrape failure (WIAA unreachable / page layout changed), never a
        # legitimate mass exodus. Capture this BEFORE the IL merge below, or
        # IL contacts would mask an empty WI scrape.
        wi_empty = bool(schools) and not admins and not coaches
        all_scraped.extend(("WI", a) for a in admins)
        all_scraped.extend(("WI", c) for c in coaches)

        # Merge IL schools into this rep's digest if configured (Andy only)
        il_count = 0
        il_scraped_schools = set()
        if rep.get("include_il") and il_schools:
            print(f"  Pulling {len(il_schools)} IL schools via IHSA API...")
            il_admins, il_coaches, il_scraped_schools = scrape_il_schools(il_schools)
            admins = admins + il_admins
            coaches = coaches + il_coaches
            il_count = len(il_schools)
            all_scraped.extend(("IL", a) for a in il_admins)
            all_scraped.extend(("IL", c) for c in il_coaches)

        # Schools that scraped successfully this run — used to scope which
        # "removed" keys we trust as real departures (see below).
        scraped_schools_this_run = wi_scraped_schools | il_scraped_schools

        current_records = contacts_to_records(admins, coaches)
        current_keys = set(current_records.keys())
        previous_keys, previous_records = load_snapshot(rep["name"])
        added, removed, first_run = diff_keys(previous_keys, current_keys)

        xlsx_bytes, sheet_summary = build_xlsx(admins, coaches, rep["name"])
        digest_label = "WI+IL" if rep.get("include_il") else "WI"
        xlsx_name = f"{rep['name'].replace(' ', '_')}-{digest_label}_School_Admins_Coaches.xlsx"

        # GUARD: don't trust a run that lost a large share of contacts with
        # nothing new added — that's a source-site scrape failure, not real
        # churn (e.g. WIAA returned empty -> the diff lists every WI contact
        # as "removed"). When suspected: do NOT overwrite the snapshot (so the
        # next run diffs against good data) and do NOT send the alarming
        # "removed everything" digest. The sheet merge is add-only and the NS
        # push keys off the Sync column, so nothing is ever deleted regardless;
        # this just stops the false alarm and the next-day phantom "+N added".
        suspect_failure = wi_empty or (
            not first_run and not added and len(removed) >= 50
            and len(current_keys) < len(previous_keys) * 0.5
        )
        if suspect_failure:
            print(f"  [GUARD] Suspected scrape failure for {rep['name']}: "
                  f"removed={len(removed)} added={len(added)} "
                  f"current={len(current_keys)} previous={len(previous_keys)} "
                  f"wi_empty={wi_empty}. Preserving snapshot; skipping digest.")
            alert = (
                f"{digest_label} digest for {rep['name']} was SKIPPED.\n\n"
                f"The scrape came back far smaller than last run "
                f"({len(removed)} contacts missing, 0 new), which means the "
                f"source site (WIAA/IHSA) was unreachable or changed its page "
                f"layout — not that contacts actually left.\n\n"
                f"No changes were made: the snapshot was preserved and no "
                f"removals were reported. Your sheet and NetSuite are untouched. "
                f"The next run will re-check automatically."
            )
            subject = f"{rep['name']} - {digest_label} digest SKIPPED (scrape looked incomplete)"
            # Only alert in shadow/dry mode (goes to Andy). In true-live mode
            # send_email would deliver to the rep, who shouldn't get a
            # "skipped" notice — just log it there. Snapshot is intentionally
            # NOT saved in either case.
            in_shadow = bool(os.environ.get("DIGESTS_OVERRIDE_TO", "").strip()) or DRY_RUN
            sent = send_email(rep, subject, alert, xlsx_bytes, xlsx_name) if in_shadow else False
            results.append({
                "rep": rep["name"], "schools": len(schools),
                "added": 0, "removed": 0, "sent": sent,
            })
            continue

        # Only trust a "removed" key as a real departure if its school
        # actually scraped successfully this run. Otherwise a single school's
        # transient WIAA fetch error (the rep-level total-loss guard above
        # doesn't catch a one-school hiccup) would look like everyone at that
        # school quit. Keys from failed-scrape schools are left untouched —
        # they'll be re-checked (and reported/acted on) the next time that
        # school's page loads successfully.
        removed_scoped = {k for k in removed if k[0] in scraped_schools_this_run}
        removed_skipped = removed - removed_scoped
        # Past the suspect-failure guard: these schools' scrapes are trusted,
        # so the end-of-run reconcile may flip their absent contacts.
        all_scraped_schools |= scraped_schools_this_run
        if removed_skipped:
            skipped_schools = sorted({k[0] for k in removed_skipped})
            print(f"  [GUARD] Ignoring {len(removed_skipped)} 'removed' key(s) at "
                  f"{len(skipped_schools)} school(s) that failed to scrape this run "
                  f"(not treated as departures): {skipped_schools[:10]}"
                  + (" ..." if len(skipped_schools) > 10 else ""))

        for k in removed_scoped:
            school, email, role, sport = k
            all_departed_triples.add((school, email, sport or role))

        def render(prefix, key, records):
            # key = (school, email, role, sport)
            rec = records.get(key, {})
            name = (f"{rec.get('first','')} {rec.get('last','')}").strip() or "(name unknown)"
            school, email, role, sport = key
            tail = f"  [{sport}]" if sport else ""
            return f"  {prefix} {name}  {email}  {role}  ({school}){tail}"

        body_lines = [f"{digest_label} school contact digest for {rep['name']}", ""]
        if first_run:
            body_lines.append("Initial snapshot — no previous version to diff against.")
        else:
            body_lines.append(f"Changes since last run: +{len(added)} / -{len(removed_scoped)}")
            if added:
                body_lines.append("\nAdded:")
                for k in sorted(added):
                    body_lines.append(render("+", k, current_records))
            if removed_scoped:
                body_lines.append("\nRemoved (marked inactive in NetSuite tonight):")
                for k in sorted(removed_scoped):
                    body_lines.append(render("-", k, previous_records))
        body_lines += ["", "Sheet counts:"] + [f"  {k}: {v}" for k, v in sorted(sheet_summary.items())]
        body = "\n".join(body_lines)

        should_send = first_run or added or removed_scoped
        if should_send:
            subject = f"{rep['name']} - Updated {digest_label} School Admins and Coaches"
            sent = send_email(rep, subject, body, xlsx_bytes, xlsx_name)
        else:
            print(f"  No changes — no email sent.")
            sent = False

        # Archive every run's XLSX to the shared Drive folder, regardless
        # of whether an email went out. Useful for auditing back in time.
        upload_digest_to_drive(rep["name"], xlsx_bytes, xlsx_name)

        # Carry forward last-known-good records for any school that failed to
        # scrape this run, so the snapshot doesn't silently lose that
        # school's people (which would falsely show them as "+N added" the
        # next time that school's page loads, and would make their real
        # absence undetectable since they'd no longer be in "previous").
        carried_records = {k: v for k, v in previous_records.items()
                           if k[0] not in scraped_schools_this_run}
        snapshot_records = {**carried_records, **current_records}
        save_snapshot(rep["name"], snapshot_records)
        results.append({
            "rep": rep["name"],
            "schools": len(schools),
            "added": len(added),
            "removed": len(removed_scoped),
            "sent": sent,
        })

    # Write the unified scrape to the master sheet's Contacts tab. Doing
    # this once at the end (after all reps) means a single Google Sheets
    # write instead of one per rep, and every scraped contact lands on
    # the sheet before the 6:30 AM NS push workflow reads it.
    merge_scraped_into_master_sheet(gc, all_scraped, all_departed_triples,
                                    scraped_schools=all_scraped_schools)

    print("\n" + "=" * 60)
    print("Summary:")
    for r in results:
        print(f"  {r['rep']:<20}  schools={r['schools']:<3}  +{r['added']:<3} -{r['removed']:<3}  sent={r['sent']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
